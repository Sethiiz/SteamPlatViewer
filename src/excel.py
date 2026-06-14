from __future__ import annotations

import os
from dataclasses import dataclass

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

OUTPUT_PATH = "platinum_list.xlsx"

# Cores por status
FILLS = {
    "Nunca jogado": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "Incompleto":   PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "Platinado":    PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
}
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=12)
CENTER      = Alignment(horizontal="center", vertical="center")
WRAP        = Alignment(horizontal="left",   vertical="center", wrap_text=True)

HEADERS    = ["#", "Jogo", "Status", "Link HLTB", "Conquistas", "Horas p/ Platinar (HLTB)"]
COL_WIDTHS = [5,   42,     16,       45,          14,           24]

# índices 1-based
COL_NUM, COL_NAME, COL_STATUS, COL_LINK, COL_ACH, COL_HOURS = 1, 2, 3, 4, 5, 6


@dataclass
class Game:
    name: str
    status: str
    link_hltb: str
    achievements: int
    hours: float


class ExcelWriter:
    def __init__(self, wb: openpyxl.Workbook) -> None:
        self._wb = wb
        self._ws = wb["Platinas"]

    @classmethod
    def new(cls) -> ExcelWriter:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Platinas"

        for col, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER
            ws.column_dimensions[cell.column_letter].width = width

        ws.row_dimensions[1].height = 20
        ws.freeze_panes = "A2"
        wb.save(OUTPUT_PATH)
        return cls(wb)

    @classmethod
    def load(cls) -> ExcelWriter | None:
        if not os.path.exists(OUTPUT_PATH):
            return None
        wb = openpyxl.load_workbook(OUTPUT_PATH)
        if "Platinas" not in wb.sheetnames:
            return None
        return cls(wb)

    def _fill(self, status: str) -> PatternFill:
        return FILLS.get(status, PatternFill())

    def append_game(self, game: Game) -> None:
        row = self._ws.max_row + 1
        fill = self._fill(game.status)

        def set_cell(col: int, value, align: Alignment = CENTER) -> None:
            cell = self._ws.cell(row=row, column=col, value=value)
            cell.fill = fill
            cell.alignment = align

        set_cell(COL_NUM,    0)
        set_cell(COL_NAME,   game.name)
        set_cell(COL_STATUS, game.status)
        set_cell(COL_LINK,   game.link_hltb, WRAP)
        set_cell(COL_ACH,    game.achievements)
        set_cell(COL_HOURS,  game.hours)
        self._wb.save(OUTPUT_PATH)

    def sort_and_number(self) -> None:
        rows = [
            r for r in self._ws.iter_rows(min_row=2, values_only=True)
            if any(v is not None for v in r)
        ]
        rows.sort(key=lambda r: r[COL_HOURS - 1] or 0)

        for row_idx, row_data in enumerate(rows, 2):
            status = row_data[COL_STATUS - 1] or ""
            fill   = self._fill(status)
            values = [row_idx - 1, row_data[1], row_data[2], row_data[3], row_data[4], row_data[5]]
            aligns = [CENTER, CENTER, CENTER, WRAP, CENTER, CENTER]
            for col, (value, align) in enumerate(zip(values, aligns), 1):
                cell = self._ws.cell(row=row_idx, column=col, value=value)
                cell.fill = fill
                cell.alignment = align

        self._wb.save(OUTPUT_PATH)

    def update_statuses(self, name_to_game: dict, steam_client, steam_id: str) -> int:
        import asyncio

        rows_to_check = [
            row for row in self._ws.iter_rows(min_row=2)
            if row[COL_STATUS - 1].value not in (None, "Platinado")
        ]
        if not rows_to_check:
            print("Todos os jogos já estão como Platinado.")
            return 0

        total = len(rows_to_check)
        print(f"\n{total} jogos não-platinados para verificar.\n")
        changed = 0

        for idx, row in enumerate(rows_to_check, 1):
            name       = row[COL_NAME - 1].value
            old_status = row[COL_STATUS - 1].value
            ach_count  = row[COL_ACH - 1].value or 0
            print(f"[{idx}/{total}] {name}...", end=" ", flush=True)

            game = name_to_game.get(name)
            if not game:
                print("não encontrado na lista Steam.")
                continue

            new_status = asyncio.get_event_loop().run_until_complete(
                steam_client.get_player_status(steam_id, game.appid, game.playtime_forever, ach_count)
            )

            if new_status == old_status:
                print(f"sem mudança ({old_status}).")
                continue

            print(f"{old_status} → {new_status} ✓")
            fill = self._fill(new_status)
            for cell in row:
                cell.fill = fill
            row[COL_STATUS - 1].value = new_status
            changed += 1

        if changed:
            self._wb.save(OUTPUT_PATH)
        return changed
